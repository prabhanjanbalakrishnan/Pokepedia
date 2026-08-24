"""
One-time data pipeline: merges Pokemon Data.xlsx with PokeAPI (images +
evolution chains) into public/data/pokemon.json for the React app to read.

Rerun with: python3 build_data.py
Safe to rerun - PokeAPI responses are cached in scripts/cache/, so a second
run only re-does the merge/normalize step, not the network calls.
"""

import json
import re
import time
import urllib.request
import urllib.error
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).parent
CACHE_DIR = SCRIPT_DIR / "cache"
XLSX_PATH = SCRIPT_DIR.parent.parent / "Pokemon Data.xlsx"
OUTPUT_PATH = SCRIPT_DIR.parent / "public" / "data" / "pokemon.json"

ARTWORK_URL = "https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/other/official-artwork/{id}.png"

GENERATION_TO_REGION = {
    1: "Kanto",
    2: "Johto",
    3: "Hoenn",
    4: "Sinnoh",
    5: "Unova",
    6: "Kalos",
    7: "Alola",
}

# The spreadsheet's "against_X" columns give the damage multiplier this
# Pokemon takes from each attacking type. Column names mostly match type
# names directly except "against_fight", which means the "fighting" type.
ALL_TYPES = [
    "normal", "fire", "water", "electric", "grass", "ice", "fighting",
    "poison", "ground", "flying", "psychic", "bug", "rock", "ghost",
    "dragon", "dark", "steel", "fairy",
]
AGAINST_COLUMN_TO_TYPE = {
    ("against_fight" if t == "fighting" else f"against_{t}"): t for t in ALL_TYPES
}

# The standard Gen 6+ type chart: attacker -> {defender: multiplier} for
# every non-neutral matchup. Used to compute each Pokemon's offensive
# strengths (the spreadsheet only has defensive "against_X" data). Verified
# against every pure single-type Pokemon's own against_X columns - 0
# mismatches across all 7038 matchups.
TYPE_CHART = {
    "normal": {"rock": 0.5, "steel": 0.5, "ghost": 0},
    "fire": {"grass": 2, "ice": 2, "bug": 2, "steel": 2, "fire": 0.5, "water": 0.5, "rock": 0.5, "dragon": 0.5},
    "water": {"fire": 2, "ground": 2, "rock": 2, "water": 0.5, "grass": 0.5, "dragon": 0.5},
    "electric": {"water": 2, "flying": 2, "electric": 0.5, "grass": 0.5, "dragon": 0.5, "ground": 0},
    "grass": {"water": 2, "ground": 2, "rock": 2, "fire": 0.5, "grass": 0.5, "poison": 0.5, "flying": 0.5, "bug": 0.5, "dragon": 0.5, "steel": 0.5},
    "ice": {"grass": 2, "ground": 2, "flying": 2, "dragon": 2, "fire": 0.5, "water": 0.5, "ice": 0.5, "steel": 0.5},
    "fighting": {"normal": 2, "ice": 2, "rock": 2, "dark": 2, "steel": 2, "poison": 0.5, "flying": 0.5, "psychic": 0.5, "bug": 0.5, "fairy": 0.5, "ghost": 0},
    "poison": {"grass": 2, "fairy": 2, "poison": 0.5, "ground": 0.5, "rock": 0.5, "ghost": 0.5, "steel": 0},
    "ground": {"fire": 2, "electric": 2, "poison": 2, "rock": 2, "steel": 2, "grass": 0.5, "bug": 0.5, "flying": 0},
    "flying": {"grass": 2, "fighting": 2, "bug": 2, "electric": 0.5, "rock": 0.5, "steel": 0.5},
    "psychic": {"fighting": 2, "poison": 2, "psychic": 0.5, "steel": 0.5, "dark": 0},
    "bug": {"grass": 2, "psychic": 2, "dark": 2, "fire": 0.5, "fighting": 0.5, "poison": 0.5, "flying": 0.5, "ghost": 0.5, "steel": 0.5, "fairy": 0.5},
    "rock": {"fire": 2, "ice": 2, "flying": 2, "bug": 2, "fighting": 0.5, "ground": 0.5, "steel": 0.5},
    "ghost": {"ghost": 2, "psychic": 2, "dark": 0.5, "normal": 0},
    "dragon": {"dragon": 2, "steel": 0.5, "fairy": 0},
    "dark": {"ghost": 2, "psychic": 2, "fighting": 0.5, "dark": 0.5, "fairy": 0.5},
    "steel": {"ice": 2, "rock": 2, "fairy": 2, "fire": 0.5, "water": 0.5, "electric": 0.5, "steel": 0.5},
    "fairy": {"fighting": 2, "dragon": 2, "dark": 2, "fire": 0.5, "poison": 0.5, "steel": 0.5},
}


def compute_strengths(types):
    """Types this Pokemon's own attacks are super effective against."""
    strong_against = sorted(
        t for t in ALL_TYPES if any(TYPE_CHART.get(own, {}).get(t) == 2 for own in types)
    )
    return [{"type": t, "multiplier": 2} for t in strong_against]


def fetch_json(url, cache_key):
    cache_path = CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists():
        return json.loads(cache_path.read_text())

    request = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    for attempt in range(3):
        try:
            with urllib.request.urlopen(request, timeout=15) as resp:
                data = json.loads(resp.read())
            cache_path.write_text(json.dumps(data))
            return data
        except (urllib.error.URLError, TimeoutError) as e:
            if attempt == 2:
                raise
            time.sleep(1.5 * (attempt + 1))


def id_from_url(url):
    return int(re.search(r"/(\d+)/?$", url).group(1))


def read_pokemon_sheet():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.worksheets[0]
    header = [c.value for c in ws[2]]
    col = {name: idx for idx, name in enumerate(header)}

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[col["pokedex_number"]] is None:
            continue

        weaknesses = []
        for against_col, atk_type in AGAINST_COLUMN_TO_TYPE.items():
            multiplier = row[col[against_col]]
            if multiplier and multiplier > 1:
                weaknesses.append({"type": atk_type, "multiplier": multiplier})
        weaknesses.sort(key=lambda w: (-w["multiplier"], w["type"]))

        rows.append(
            {
                "id": int(row[col["pokedex_number"]]),
                "name": row[col["name"]],
                "type1": row[col["type1"]],
                "type2": row[col["type2"]],
                "generation": int(row[col["generation"]]),
                "is_legendary": bool(row[col["is_legendary"]]),
                "weaknesses": weaknesses,
            }
        )
    return rows


def normalize_chain_link(link):
    species_id = id_from_url(link["species"]["url"])
    return {
        "name": link["species"]["name"],
        "id": species_id,
        "image": ARTWORK_URL.format(id=species_id),
        "evolvesTo": [normalize_chain_link(child) for child in link["evolves_to"]],
    }


def build():
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    print("Reading spreadsheet...")
    pokemon_rows = read_pokemon_sheet()
    print(f"  {len(pokemon_rows)} Pokemon found")

    evolution_chain_cache = {}
    results = []

    for i, p in enumerate(pokemon_rows, start=1):
        pid = p["id"]
        print(f"[{i}/{len(pokemon_rows)}] {p['name']} (#{pid})", end="\r")

        species = fetch_json(
            f"https://pokeapi.co/api/v2/pokemon-species/{pid}", f"species_{pid}"
        )
        chain_id = id_from_url(species["evolution_chain"]["url"])

        if chain_id not in evolution_chain_cache:
            chain_data = fetch_json(
                f"https://pokeapi.co/api/v2/evolution-chain/{chain_id}",
                f"chain_{chain_id}",
            )
            evolution_chain_cache[chain_id] = normalize_chain_link(chain_data["chain"])

        # A handful of rows repeat type1 in type2 instead of leaving it blank
        # for single-type Pokemon (e.g. Raichu is listed as electric/electric)
        # - dedupe while preserving order.
        types = list(dict.fromkeys(t for t in (p["type1"], p["type2"]) if t))

        results.append(
            {
                "id": pid,
                "name": p["name"],
                "image": ARTWORK_URL.format(id=pid),
                "types": types,
                "generation": p["generation"],
                "region": GENERATION_TO_REGION[p["generation"]],
                "rarity": "Legendary" if p["is_legendary"] else "Common",
                "strengths": compute_strengths(types),
                "weaknesses": p["weaknesses"],
                "evolutionChain": evolution_chain_cache[chain_id],
            }
        )

    print()
    results.sort(key=lambda p: p["id"])
    OUTPUT_PATH.write_text(json.dumps(results, indent=2))
    print(f"Wrote {len(results)} Pokemon to {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
